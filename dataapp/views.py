from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count
from .models import ExcelData, Friend
from .forms import ExcelDataForm, FriendForm, RegisterForm

# Authentication Views
def login_view(request):
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'dataapp/login.html')

def register_view(request):
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Account created successfully!')
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = RegisterForm()
    
    return render(request, 'dataapp/register.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully')
    return redirect('login')

# Dashboard Views
@login_required
def admin_dashboard(request):
    total_records = ExcelData.objects.filter(created_by=request.user).count()
    total_friends = Friend.objects.filter(user=request.user).count()
    recent_data = ExcelData.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    recent_friends = Friend.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    context = {
        'total_records': total_records,
        'total_friends': total_friends,
        'recent_data': recent_data,
        'recent_friends': recent_friends,
    }
    return render(request, 'dataapp/admin_dashboard.html', context)

# Excel Data Views
@login_required
def data_list_view(request):
    data = ExcelData.objects.filter(created_by=request.user)
    search = request.GET.get('search', '')
    
    if search:
        data = data.filter(column1__icontains=search) | data.filter(column2__icontains=search)
    
    paginator = Paginator(data, 20)
    page = request.GET.get('page')
    data_page = paginator.get_page(page)
    
    context = {
        'data_page': data_page,
        'search': search,
    }
    return render(request, 'dataapp/data_list_admin.html', context)

@login_required
def data_add_view(request):
    if request.method == 'POST':
        form = ExcelDataForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = request.user
            data.save()
            messages.success(request, 'Record added successfully!')
            return redirect('data_list_admin')
    else:
        form = ExcelDataForm()
    
    return render(request, 'dataapp/data_form.html', {'form': form, 'mode': 'add'})

@login_required
def data_edit_view(request, pk):
    data = get_object_or_404(ExcelData, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = ExcelDataForm(request.POST, instance=data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Record updated successfully!')
            return redirect('data_list_admin')
    else:
        form = ExcelDataForm(instance=data)
    
    return render(request, 'dataapp/data_form.html', {'form': form, 'mode': 'edit', 'record': data})

@login_required
def data_delete_view(request, pk):
    data = get_object_or_404(ExcelData, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        data.delete()
        messages.success(request, 'Record deleted successfully!')
        return redirect('data_list_admin')
    
    return render(request, 'dataapp/data_delete.html', {'record': data})

# Friends Views
@login_required
def friends_list_view(request):
    friends = Friend.objects.filter(user=request.user)
    search = request.GET.get('search', '')
    
    if search:
        friends = friends.filter(name__icontains=search) | friends.filter(email__icontains=search)
    
    paginator = Paginator(friends, 20)
    page = request.GET.get('page')
    friends_page = paginator.get_page(page)
    
    context = {
        'friends_page': friends_page,
        'search': search,
    }
    return render(request, 'dataapp/friends_list.html', context)

@login_required
def friend_add_view(request):
    if request.method == 'POST':
        form = FriendForm(request.POST)
        if form.is_valid():
            friend = form.save(commit=False)
            friend.user = request.user
            friend.save()
            messages.success(request, 'Friend added successfully!')
            return redirect('friends_list')
    else:
        form = FriendForm()
    
    return render(request, 'dataapp/friend_form.html', {'form': form, 'mode': 'add'})

@login_required
def friend_edit_view(request, pk):
    friend = get_object_or_404(Friend, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = FriendForm(request.POST, instance=friend)
        if form.is_valid():
            form.save()
            messages.success(request, 'Friend updated successfully!')
            return redirect('friends_list')
    else:
        form = FriendForm(instance=friend)
    
    return render(request, 'dataapp/friend_form.html', {'form': form, 'mode': 'edit', 'friend': friend})

@login_required
def friend_delete_view(request, pk):
    friend = get_object_or_404(Friend, pk=pk, user=request.user)
    
    if request.method == 'POST':
        friend.delete()
        messages.success(request, 'Friend deleted successfully!')
        return redirect('friends_list')
    
    return render(request, 'dataapp/friend_delete.html', {'friend': friend})

@login_required
def export_to_excel(request):
    """Export Excel data to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Excel Data"
    
    # Add headers
    headers = ['ID', 'Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 5', 'Column 6', 'Created At']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    data = ExcelData.objects.filter(created_by=request.user)
    for item in data:
        ws.append([
            item.id,
            item.column1,
            item.column2,
            item.column3,
            item.column4,
            item.column5,
            item.column6.strftime('%Y-%m-%d') if item.column6 else '',
            item.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    now_myt = datetime.now(malaysia_tz)
    filename = f'excel_data_{now_myt.strftime("%Y%m%d_%H%M%S")}_MYT.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response